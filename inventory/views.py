from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.db.models import Count, Sum, Avg, F, Q
from django.db import IntegrityError
from .models import NWO, TelephoneExchange, Cable, Equipment, EBCircuit, MobileBTS, JunctionBox, LIU, Fiber, Splicing, FTTH, OHMaintenanceEntry, OHMaintenanceActivity, OHMaintenanceRateMaster
from .forms import LIUForm, JBForm, CableForm, EquipmentForm, CircuitForm, BTSForm, FTTHForm, ChangePasswordForm, OHMaintenanceEntryForm, OHMaintenanceActivityFormSet
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone
import logging
# import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
import json
from django.views.decorators.http import require_http_methods

class DivisionRequiredMixin:
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_superuser:
            if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
                division = self.request.user.profile.division
                # Filter based on the model type
                if self.model == Cable:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == Equipment:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == EBCircuit:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == MobileBTS:
                    queryset = queryset.filter(maan_node__te__nwo=division)
                elif self.model == JunctionBox:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == LIU:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == FTTH:
                    queryset = queryset.filter(division=division)
                elif self.model == OHMaintenanceEntry:
                    queryset = queryset.filter(division=division)
        return queryset

class TEContextMixin:
    def get_initial(self):
        initial = super().get_initial()
        te_id = self.request.GET.get('te')
        if te_id:
            initial['te'] = te_id
        return initial


from .defaults import DEFAULT_PASSWORDS


def _get_division_default_password(division_name):
    return DEFAULT_PASSWORDS.get(division_name)


@login_required
@require_http_methods(["POST"])
def reset_password(request):
    profile = getattr(request.user, 'profile', None)
    if not profile or not profile.division:
        messages.error(request, 'Unable to reset password: division is not configured for your account.')
        return redirect('change_password')

    default_password = _get_division_default_password(profile.division.name)
    if not default_password:
        messages.error(request, 'Unable to reset password: default password is not configured for your division.')
        return redirect('change_password')

    request.user.set_password(default_password)
    request.user.save()

    profile.force_password_change = True
    profile.last_password_reset = timezone.now()
    profile.save(update_fields=['force_password_change', 'last_password_reset'])

    logger = logging.getLogger(__name__)
    logger.info('Password reset via division default for user=%s division=%s', request.user.username, profile.division.name)

    logout(request)
    messages.success(request, 'Password reset successfully. Please login using the default division password and change it immediately.')
    return redirect('login')


@login_required
@require_http_methods(["GET", "POST"])
def change_password(request):
    if request.method == "GET":
        form = ChangePasswordForm(user=request.user)
        force_change = bool(getattr(request.user, 'profile', None) and request.user.profile.force_password_change)
        return render(request, "inventory/change_password.html", {"form": form, "force_password_change": force_change})

    is_json = "application/json" in (request.content_type or "")
    if is_json:
        try:
            payload = json.loads((request.body or b"{}").decode("utf-8"))
        except Exception:
            return HttpResponseBadRequest("Invalid JSON")
        if "new_password" in payload and "confirm_new_password" not in payload:
            payload["confirm_new_password"] = payload.get("new_password")
        form = ChangePasswordForm(payload, user=request.user)
    else:
        form = ChangePasswordForm(request.POST, user=request.user)

    if form.is_valid():
        request.user.set_password(form.cleaned_data["new_password"])
        request.user.save()
        profile = getattr(request.user, 'profile', None)
        if profile and profile.force_password_change:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])
        logout(request)
        msg = "Password changed successfully. Please login again."
        if is_json:
            return JsonResponse({"detail": msg})
        messages.success(request, msg)
        return redirect("login")

    if is_json:
        return JsonResponse({"errors": form.errors}, status=400)
    force_change = bool(getattr(request.user, 'profile', None) and request.user.profile.force_password_change)
    return render(request, "inventory/change_password.html", {"form": form, "force_password_change": force_change})


@login_required
def te_jb_routes(request, te_id):
    te = get_object_or_404(TelephoneExchange, id=te_id)
    if not request.user.is_superuser and hasattr(request.user, "profile") and getattr(request.user.profile, "division_id", None):
        if te.nwo_id != request.user.profile.division_id:
            return HttpResponse(status=403)

    jbs = list(
        JunctionBox.objects.filter(te=te).prefetch_related("input_cables", "output_cables")
    )

    def build_routes(mode):
        nodes = [jb for jb in jbs if jb.cable_mode == mode]
        if not nodes:
            return []

        inputs = {jb.id: set(c.id for c in jb.input_cables.all()) for jb in nodes}
        outputs = {jb.id: set(c.id for c in jb.output_cables.all()) for jb in nodes}
        edges = {jb.id: [] for jb in nodes}
        incoming = {jb.id: 0 for jb in nodes}

        for a in nodes:
            out_set = outputs.get(a.id, set())
            if not out_set:
                continue
            for b in nodes:
                if a.id == b.id:
                    continue
                if out_set & inputs.get(b.id, set()):
                    edges[a.id].append(b.id)
                    incoming[b.id] += 1

        starts = [jb_id for jb_id, deg in incoming.items() if deg == 0]
        if not starts:
            starts = [jb.id for jb in nodes]

        id_to = {jb.id: jb for jb in nodes}
        routes = []

        def dfs(cur, path, seen):
            if len(routes) >= 100:
                return
            nxts = edges.get(cur, [])
            if not nxts:
                routes.append([id_to[i] for i in path])
                return
            extended = False
            for nxt in nxts:
                if nxt in seen:
                    continue
                extended = True
                dfs(nxt, path + [nxt], seen | {nxt})
            if not extended:
                routes.append([id_to[i] for i in path])

        for s in starts:
            dfs(s, [s], {s})

        uniq = []
        seen_keys = set()
        for r in routes:
            key = tuple(j.id for j in r)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            uniq.append(r)
        return uniq

    ug_routes = build_routes("UG")
    oh_routes = build_routes("OH")

    def points(mode):
        pts = []
        for jb in jbs:
            if jb.cable_mode != mode:
                continue
            if jb.latitude is None or jb.longitude is None:
                continue
            pts.append(
                {
                    "id": jb.id,
                    "jb_id": jb.jb_id,
                    "name": jb.jb_name or "",
                    "lat": float(jb.latitude),
                    "lng": float(jb.longitude),
                    "url": reverse("jb_detail", args=[jb.id]),
                }
            )
        return pts

    return render(
        request,
        "inventory/te_jb_routes.html",
        {
            "te": te,
            "ug_routes": ug_routes,
            "oh_routes": oh_routes,
            "ug_points_json": json.dumps(points("UG")),
            "oh_points_json": json.dumps(points("OH")),
        },
    )

@login_required
def dashboard(request):
    # 1. Identify User Division
    user_division = None
    if hasattr(request.user, 'profile') and request.user.profile.division:
        user_division = request.user.profile.division

    # 2. Fetch Divisions
    if user_division:
        nwos_to_process = NWO.objects.filter(id=user_division.id)
    else:
        nwos_to_process = NWO.objects.all()

    # 3. Sort according to preference
    nwo_order = [
        'NWO CENTRAL', 'NWO PALARIVATTOM', 'NWO KOCHI', 'NWO TRIPUNITHARA',
        'NWO ANGAMALY', 'NWO THODUPUZHA', 'NWO ALUVA', 'NWO MOOVATTUPUZHA',
        'NWO ADIMALY', 'NWO KATTAPPANA'
    ]
    
    nwo_map = {n.name: n for n in nwos_to_process}
    sorted_nwos = []
    for code in nwo_order:
        if code in nwo_map:
            sorted_nwos.append(nwo_map[code])
    
    # Add any other divisions that might not be in the order list
    for nwo in nwos_to_process:
        if nwo not in sorted_nwos:
            sorted_nwos.append(nwo)

    # 4. Build data structure
    nwo_data_list = []
    for nwo in sorted_nwos:
        exchanges = nwo.exchanges.all().order_by('name')
        te_list = []
        for te in exchanges:
            te_list.append({
                'id': te.id,
                'name': te.name,
                'cable_count': te.cables.count(),
                'equipment_count': te.equipments.count(),
                'circuit_count': EBCircuit.objects.filter(te=te).count(),
                'bts_count': MobileBTS.objects.filter(maan_node__te=te).count(),
                'ftth_count': FTTH.objects.filter(te=te).count(),
            })
        
        nwo_data_list.append({
            'display_name': nwo.name,
            'te_data': te_list
        })

    # 5. Global Stats
    if user_division:
        total_tes = TelephoneExchange.objects.filter(nwo=user_division).count()
        total_circuits = EBCircuit.objects.filter(te__nwo=user_division).count()
        total_cables = Cable.objects.filter(te__nwo=user_division).count()
        total_equipment = Equipment.objects.filter(te__nwo=user_division).count()
        # For BTS, we count linked ones. For FTTH, we use the direct division relation.
        total_bts = MobileBTS.objects.filter(maan_node__te__nwo=user_division).count()
        total_ftth = FTTH.objects.filter(division=user_division).count()
    else:
        total_tes = TelephoneExchange.objects.count()
        total_circuits = EBCircuit.objects.count()
        total_cables = Cable.objects.count()
        total_equipment = Equipment.objects.count()
        total_bts = MobileBTS.objects.count()
        total_ftth = FTTH.objects.count()

    context = {
        'nwo_data_list': nwo_data_list,
        'total_tes': total_tes,
        'total_circuits': total_circuits,
        'total_cables': total_cables,
        'total_equipment': total_equipment,
        'total_bts': total_bts,
        'total_ftth': total_ftth,
        'is_filtered': user_division is not None,
        'user_division_name': user_division.name if user_division else "All Divisions",
    }
    return render(request, 'inventory/dashboard.html', context)

@login_required
def te_dashboard(request, te_id):
    te = get_object_or_404(TelephoneExchange, id=te_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if te.nwo != request.user.profile.division:
                messages.error(request, "You do not have permission to access this TE.")
                return redirect('dashboard')
    
    cables = Cable.objects.filter(te=te)
    equipments = Equipment.objects.filter(te=te)
    bts_count = MobileBTS.objects.filter(maan_node__te=te).count()
    
    # Cable summary
    ug_summary = cables.filter(mode='UG').values('cable_type').annotate(count=Count('id')).order_by('cable_type')
    oh_summary = cables.filter(mode='OH').values('cable_type').annotate(count=Count('id')).order_by('cable_type')
    
    ug_dict = {item['cable_type']: item['count'] for item in ug_summary}
    oh_dict = {item['cable_type']: item['count'] for item in oh_summary}
    
    # Cable categories
    in_cables = cables.filter(category='IN').count()
    out_cables = cables.filter(category='OUT').count()
    tie_cables = cables.filter(category='TIE').count()
    
    # Convert Decimals to floats for JS compatibility in JBs
    jbs_raw = JunctionBox.objects.filter(te=te)
    jbs = []
    for jb in jbs_raw:
        jbs.append({
            'jb_id': jb.jb_id,
            'latitude': float(jb.latitude) if jb.latitude else None,
            'longitude': float(jb.longitude) if jb.longitude else None,
        })
    
    bts_list = MobileBTS.objects.filter(maan_node__te=te)
    
    # Fetch route points for cables in this TE
    cable_routes = []
    for cable in cables:
        # Convert Decimals to floats for JS compatibility
        points = [[float(p[0]), float(p[1])] for p in cable.route_points.all().values_list('latitude', 'longitude', flat=False)]
        if points:
            cable_routes.append({
                'name': cable.name,
                'type': cable.cable_type,
                'mode': cable.mode,
                'points': points
            })
    
    context = {
        'te': te,
        'cables': cables,
        'equipments': equipments,
        'bts_count': bts_count,
        'bts_list': bts_list,
        'ug_summary': ug_dict,
        'oh_summary': oh_dict,
        'in_cables': in_cables,
        'out_cables': out_cables,
        'tie_cables': tie_cables,
        'jbs': jbs,
        'cable_routes': cable_routes,
    }
    return render(request, 'inventory/te_dashboard.html', context)

class CableListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = Cable
    template_name = 'inventory/cable_list.html'
    context_object_name = 'cables'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('te', 'te__nwo')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        # Cable type filter
        cable_type = self.request.GET.get('cable_type')
        if cable_type:
            queryset = queryset.filter(cable_type=cable_type)
        
        # Mode filter (OH/UG)
        mode = self.request.GET.get('mode')
        if mode:
            queryset = queryset.filter(mode=mode)
        
        # Category filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_term'] = self.request.GET.get('search', '')
        return context

class EquipmentListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = Equipment
    template_name = 'inventory/equipment_list.html'
    context_object_name = 'equipments'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('te', 'te__nwo')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        # Equipment type filter
        eq_type = self.request.GET.get('equipment_type')
        if eq_type:
            queryset = queryset.filter(equipment_type=eq_type)
        
        # TE filter
        te_id = self.request.GET.get('te')
        if te_id:
            queryset = queryset.filter(te_id=te_id)
        
        return queryset

class CircuitListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = EBCircuit
    template_name = 'inventory/circuit_list.html'
    context_object_name = 'circuits'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('te', 'te__nwo', 'cable', 'equipment')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(client_name__icontains=search)
        
        # Circuit type filter
        circuit_type = self.request.GET.get('circuit_type')
        if circuit_type:
            queryset = queryset.filter(circuit_type=circuit_type)
        
        # TE filter
        te_id = self.request.GET.get('te')
        if te_id:
            queryset = queryset.filter(te_id=te_id)
        
        return queryset

class BTSListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = MobileBTS
    template_name = 'inventory/bts_list.html'
    context_object_name = 'bts_list'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('maan_node', 'maan_node__te')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(rp_id__icontains=search) | 
                Q(bts_name__icontains=search)
            )
        
        # Ring filter
        is_ring = self.request.GET.get('is_ring')
        if is_ring:
            queryset = queryset.filter(is_ring=is_ring == 'true')
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context['total_count'] = queryset.count()
        context['ring_count'] = queryset.filter(is_ring=True).count()
        context['cef_count'] = queryset.filter(has_cef_12t=True).count()
        context['avg_power'] = queryset.aggregate(Avg('receive_power_db'))['receive_power_db__avg']
        return context

@login_required
def te_liu_setup(request, te_id):
    te = get_object_or_404(TelephoneExchange, id=te_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if te.nwo != request.user.profile.division:
                messages.error(request, "You do not have permission to access this TE.")
                return redirect('dashboard')

    cables = Cable.objects.filter(te=te)
    
    if request.method == 'POST':
        liu_count = int(request.POST.get('liu_count', 0))
        for i in range(1, liu_count + 1):
            name = request.POST.get(f'liu_{i}_name')
            cable_id = request.POST.get(f'liu_{i}_cable')
            cable_manual = request.POST.get(f'liu_{i}_cable_manual')
            capacity = int(request.POST.get(f'liu_{i}_capacity', 0))
            remarks = request.POST.get(f'liu_{i}_remarks')
            
            if name and capacity and (cable_id or cable_manual):
                cable = None
                cable_manual_entry = None
                
                if cable_id:
                    # Use selected cable from dropdown
                    cable = get_object_or_404(Cable, id=cable_id)
                else:
                    # Use manual cable entry
                    cable_manual_entry = cable_manual
                
                liu = LIU.objects.create(
                    te=te,
                    name=name,
                    cable=cable,
                    cable_manual_entry=cable_manual_entry,
                    capacity=capacity,
                    remarks=remarks or f"Setup for {te.name}"
                )
                
                # Update ports with fiber/circuit data
                for port_num in range(1, capacity + 1):
                    circuit = request.POST.get(f'liu_{i}_fiber_{port_num}_circuit')
                    sys_port = request.POST.get(f'liu_{i}_fiber_{port_num}_port')
                    otdr_distance = request.POST.get(f'liu_{i}_fiber_{port_num}_otdr_distance')
                    otdr_image = request.FILES.get(f'liu_{i}_fiber_{port_num}_otdr_image')
                    
                    if circuit or sys_port or otdr_distance or otdr_image:
                        port = liu.ports.get(port_number=port_num)
                        port.connected_to = sys_port
                        port.remarks = circuit
                        port.otdr_distance = otdr_distance
                        if otdr_image:
                            port.otdr_image = otdr_image
                        if circuit or sys_port:
                            port.status = 'Used'
                        port.save()
        
        messages.success(request, f"Successfully set up {liu_count} LIUs for {te.name}. Please configure fiber details for each LIU below.")
        return redirect(f"{reverse_lazy('liu_list')}?te={te.id}")

    return render(request, 'inventory/te_liu_setup.html', {
        'te': te,
        'cables': cables,
        'range_10': range(1, 11)
    })


@login_required
def get_cable_details(request, cable_id):
    try:
        cable = Cable.objects.get(id=cable_id)
        if not request.user.is_superuser:
            if hasattr(request.user, 'profile') and request.user.profile.division:
                if cable.te.nwo != request.user.profile.division:
                    return JsonResponse({'error': 'Access denied'}, status=403)
        
        return JsonResponse({
            'id': cable.id,
            'name': cable.name,
            'cable_type': cable.cable_type,
            'fiber_count': cable.fiber_count,
            'structure_type': cable.structure_type,
            'mode': cable.get_mode_display(),
            'category': cable.get_category_display(),
            'te': cable.te.name,
            'connected_te': cable.connected_te.name if cable.connected_te else 'N/A',
            'otdr_distance': cable.otdr_distance_formatted or 'Not recorded',
            'otdr_image_url': cable.otdr_image.url if cable.otdr_image else None,
            'remarks': cable.remarks,
        })
    except Cable.DoesNotExist:
        return JsonResponse({'error': 'Cable not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


class LIUListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = LIU
    template_name = 'inventory/liu_list.html'
    context_object_name = 'liu_list'

    def get_queryset(self):
        queryset = super().get_queryset()
        te_id = self.request.GET.get('te')
        if te_id:
            queryset = queryset.filter(te_id=te_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        te_id = self.request.GET.get('te')
        if te_id:
            context['filter_te'] = get_object_or_404(TelephoneExchange, id=te_id)
        return context

class LIUCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CreateView):
    model = LIU
    form_class = LIUForm
    template_name = 'inventory/liu_form.html'
    success_url = reverse_lazy('liu_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        liu = self.object
        capacity = liu.capacity
        
        # Update ports with fiber/circuit data from the form
        for port_num in range(1, capacity + 1):
            circuit = self.request.POST.get(f'fiber_{port_num}_circuit')
            sys_port = self.request.POST.get(f'fiber_{port_num}_port')
            
            if circuit or sys_port:
                port = liu.ports.get(port_number=port_num)
                port.connected_to = sys_port
                port.remarks = circuit
                port.status = 'Used'
                port.save()
        
        messages.success(self.request, f"LIU {liu.name} created successfully with fiber details.")
        return response

@login_required
def liu_detail(request, liu_id):
    liu = get_object_or_404(LIU, id=liu_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if liu.te.nwo != request.user.profile.division:
                messages.error(request, "You do not have permission to access this LIU.")
                return redirect('dashboard')

    ports = liu.ports.all()
    fibers = liu.cable.fibers.all() if liu.cable else Fiber.objects.none()

    if request.method == 'POST':
        for port in ports:
            fiber_id = request.POST.get(f'port_{port.id}_fiber')
            connected_to = request.POST.get(f'port_{port.id}_connected')
            status = request.POST.get(f'port_{port.id}_status')
            remarks = request.POST.get(f'port_{port.id}_remarks')

            if fiber_id:
                try:
                    port.fiber = Fiber.objects.get(id=fiber_id)
                except Fiber.DoesNotExist:
                    port.fiber = None
            else:
                port.fiber = None
            
            port.connected_to = connected_to
            port.status = status
            port.remarks = remarks
            port.save()
        
        messages.success(request, "LIU Ports updated successfully!")
        return redirect('liu_detail', liu_id=liu.id)

    return render(request, 'inventory/liu_detail.html', {
        'liu': liu,
        'ports': ports,
        'fibers': fibers
    })

class LIUUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = LIU
    form_class = LIUForm
    template_name = 'inventory/liu_form.html'
    
    def get_success_url(self):
        return reverse_lazy('liu_detail', kwargs={'liu_id': self.object.id})

    def form_valid(self, form):
        messages.success(self.request, f"LIU {self.object.name} updated successfully.")
        return super().form_valid(form)

class JBListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = JunctionBox
    template_name = 'inventory/jb_list.html'
    context_object_name = 'jbs'

class JBCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CreateView):
    model = JunctionBox
    form_class = JBForm
    template_name = 'inventory/jb_form.html'
    success_url = reverse_lazy('jb_list')

class JBUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = JunctionBox
    form_class = JBForm
    template_name = 'inventory/jb_form.html'
    success_url = reverse_lazy('jb_list')

class JBDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = JunctionBox
    template_name = 'inventory/jb_confirm_delete.html'
    success_url = reverse_lazy('jb_list')

class JBDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = JunctionBox
    template_name = 'inventory/jb_detail.html'
    context_object_name = 'jb'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['splices'] = self.object.splices.all()
        return context

@login_required
def jb_splicing(request, jb_id):
    jb = get_object_or_404(JunctionBox, id=jb_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if jb.te.nwo != request.user.profile.division:
                messages.error(request, "You do not have permission to access this Junction Box.")
                return redirect('dashboard')

    splices = jb.splices.all()
    cables = Cable.objects.filter(te=jb.te)
    
    if request.method == 'POST':
        fiber_in_id = request.POST.get('fiber_in')
        fiber_out_id = request.POST.get('fiber_out')
        remarks = request.POST.get('remarks')
        
        if fiber_in_id and fiber_out_id:
            fiber_in = get_object_or_404(Fiber, id=fiber_in_id)
            fiber_out = get_object_or_404(Fiber, id=fiber_out_id)
            
            Splicing.objects.create(
                jb=jb,
                fiber_in=fiber_in,
                fiber_out=fiber_out,
                remarks=remarks
            )
            messages.success(request, "Splice recorded successfully!")
            return redirect('jb_splicing', jb_id=jb.id)
            
    return render(request, 'inventory/jb_splicing.html', {
        'jb': jb,
        'splices': splices,
        'cables': cables
    })

@login_required
def api_get_fibers(request, cable_id):
    cable = get_object_or_404(Cable, id=cable_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if cable.te.nwo != request.user.profile.division:
                return JsonResponse({'error': 'Permission denied'}, status=403)

    fibers = cable.fibers.all().values('id', 'fiber_number', 'status', 'is_used')
    return JsonResponse({'fibers': list(fibers)})

class BTSCreateView(LoginRequiredMixin, DivisionRequiredMixin, CreateView):
    model = MobileBTS
    form_class = BTSForm
    template_name = 'inventory/bts_form.html'
    success_url = reverse_lazy('bts_list')

class BTSUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = MobileBTS
    form_class = BTSForm
    template_name = 'inventory/bts_form.html'
    success_url = reverse_lazy('bts_list')

class BTSDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = MobileBTS
    template_name = 'inventory/bts_confirm_delete.html'
    success_url = reverse_lazy('bts_list')

# FTTH Views
class FTTHListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = FTTH
    template_name = 'inventory/ftth_list.html'
    context_object_name = 'ftth_list'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('division', 'te')
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(customer_name__icontains=search) |
                Q(landline_number__icontains=search) |
                Q(olt_name__icontains=search)
            )
        return queryset

class FTTHDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = FTTH
    template_name = 'inventory/ftth_detail.html'
    context_object_name = 'ftth'

class FTTHCreateView(LoginRequiredMixin, DivisionRequiredMixin, CreateView):
    model = FTTH
    form_class = FTTHForm
    template_name = 'inventory/ftth_form.html'
    success_url = reverse_lazy('ftth_list')

    def get_initial(self):
        initial = super().get_initial()
        if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
            initial['division'] = self.request.user.profile.division
        
        te_id = self.request.GET.get('te')
        if te_id:
            initial['te'] = te_id
        return initial

class FTTHUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = FTTH
    form_class = FTTHForm
    template_name = 'inventory/ftth_form.html'
    success_url = reverse_lazy('ftth_list')

class FTTHDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = FTTH
    template_name = 'inventory/ftth_confirm_delete.html'
    success_url = reverse_lazy('ftth_list')

class BTSDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = MobileBTS
    template_name = 'inventory/bts_detail.html'
    context_object_name = 'bts'

import re

class CableFiberUpdateMixin:
    def form_valid(self, form):
        response = super().form_valid(form)
        cable = self.object
        for i in range(1, cable.fiber_count + 1):
            sys_end = self.request.POST.get(f'fiber_{i}_system_end')
            circ_name = self.request.POST.get(f'fiber_{i}_circuit_name')
            dist_str = self.request.POST.get(f'fiber_{i}_otdr_distance')
            img = self.request.FILES.get(f'fiber_{i}_otdr_image')
            
            try:
                fiber = cable.fibers.get(fiber_number=i)
            except Fiber.DoesNotExist:
                continue
                
            updated = False
            if sys_end is not None:
                fiber.system_end = sys_end
                updated = True
            if circ_name is not None:
                fiber.circuit_name = circ_name
                updated = True
            if dist_str:
                match = re.search(r'([\d.]+)', dist_str)
                if match:
                    try:
                        val = float(match.group(1))
                        if 'km' in dist_str.lower():
                            val = val * 1000
                        fiber.otdr_distance = val
                        updated = True
                    except ValueError:
                        pass
            if img:
                fiber.otdr_image = img
                updated = True
                
            if sys_end or circ_name:
                fiber.is_used = True
                fiber.status = 'Used'
                
            if updated:
                fiber.save()
        return response

class CableCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CableFiberUpdateMixin, CreateView):
    model = Cable
    form_class = CableForm
    template_name = 'inventory/cable_form.html'
    success_url = reverse_lazy('cable_list')

class CableUpdateView(LoginRequiredMixin, DivisionRequiredMixin, CableFiberUpdateMixin, UpdateView):
    model = Cable
    form_class = CableForm
    template_name = 'inventory/cable_form.html'
    success_url = reverse_lazy('cable_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fibers'] = self.object.fibers.all()
        return context

class CableDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = Cable
    template_name = 'inventory/cable_confirm_delete.html'
    success_url = reverse_lazy('cable_list')

class CableDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = Cable
    template_name = 'inventory/cable_detail.html'
    context_object_name = 'cable'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fibers'] = self.object.fibers.all()
        return context

class EquipmentCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CreateView):
    model = Equipment
    form_class = EquipmentForm
    template_name = 'inventory/equipment_form.html'
    success_url = reverse_lazy('equipment_list')

    def get_initial(self):
        initial = super().get_initial()
        eq_type = self.request.GET.get('type')
        if eq_type:
            initial['equipment_type'] = eq_type
        return initial

class EquipmentUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = Equipment
    form_class = EquipmentForm
    template_name = 'inventory/equipment_form.html'
    success_url = reverse_lazy('equipment_list')

class EquipmentDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = Equipment
    template_name = 'inventory/equipment_confirm_delete.html'
    success_url = reverse_lazy('equipment_list')

class EquipmentDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = Equipment
    template_name = 'inventory/equipment_detail.html'
    context_object_name = 'equipment'

class CircuitCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CreateView):
    model = EBCircuit
    form_class = CircuitForm
    template_name = 'inventory/circuit_form.html'
    success_url = reverse_lazy('circuit_list')

class CircuitUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = EBCircuit
    form_class = CircuitForm
    template_name = 'inventory/circuit_form.html'
    success_url = reverse_lazy('circuit_list')

class CircuitDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = EBCircuit
    template_name = 'inventory/circuit_confirm_delete.html'
    success_url = reverse_lazy('circuit_list')

class CircuitDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = EBCircuit
    template_name = 'inventory/circuit_detail.html'
    context_object_name = 'circuit'

@login_required
def analytics(request):
    from django.db.models import F, ExpressionWrapper, DurationField, Avg, Count, Sum, Q
    from django.utils import timezone
    
    division = None
    if request.user.is_superuser:
        division_id = request.GET.get('division')
        if division_id:
            division = NWO.objects.filter(id=division_id).first()
    else:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            division = request.user.profile.division

    # Fault Tracking Analytics
    fault_entries = OHMaintenanceEntry.objects.filter(fault_occurrence_datetime__isnull=False)
    if division:
        fault_entries = fault_entries.filter(division=division)
    
    # 1. Fault Duration Analysis (Closed Faults)
    closed_faults = fault_entries.filter(fault_clearance_datetime__isnull=False).annotate(
        duration=ExpressionWrapper(
            F('fault_clearance_datetime') - F('fault_occurrence_datetime'),
            output_field=DurationField()
        )
    )
    
    # 2. Division-wise MTTR
    division_mttr = closed_faults.values('division__name').annotate(
        avg_mttr=Avg('duration'),
        total_closed=Count('id')
    ).order_by('avg_mttr')

    # 3. Team-wise Performance
    team_performance = closed_faults.values('team_name', 'division__name').annotate(
        avg_mttr=Avg('duration'),
        total_closed=Count('id')
    ).order_by('avg_mttr')

    # 4. Repeated Fault Locations
    repeated_locations = fault_entries.values('location', 'division__name').annotate(
        fault_count=Count('id')
    ).filter(fault_count__gt=1).order_by('-fault_count')

    # 5. Fault Type Trends
    fault_type_trends = fault_entries.values('fault_type').annotate(
        count=Count('id')
    ).order_by('-count')

    # 6. Severity Distribution
    severity_dist = fault_entries.values('fault_severity').annotate(
        count=Count('id')
    ).order_by('-count')

    # Existing Analytics...
    # Mobile Tower RP ID Summary
    bts_list = MobileBTS.objects.select_related('maan_node', 'maan_node__te').all()
    if division:
        bts_list = bts_list.filter(maan_node__te__nwo=division)
    
    # Cable Summary - OH and UG separately
    oh_cables = Cable.objects.filter(mode='OH').annotate(
        total_fibers=Count('fibers'),
        used_fibers=Count('fibers', filter=Q(fibers__is_used=True)),
        free_fibers=Count('fibers', filter=Q(fibers__is_used=False))
    ).select_related('te')
    
    ug_cables = Cable.objects.filter(mode='UG').annotate(
        total_fibers=Count('fibers'),
        used_fibers=Count('fibers', filter=Q(fibers__is_used=True)),
        free_fibers=Count('fibers', filter=Q(fibers__is_used=False))
    ).select_related('te')

    if division:
        oh_cables = oh_cables.filter(te__nwo=division)
        ug_cables = ug_cables.filter(te__nwo=division)
    
    # Overall cable summary
    cable_queryset = Cable.objects.all()
    if division:
        cable_queryset = cable_queryset.filter(te__nwo=division)
    cable_summary = cable_queryset.values('cable_type', 'mode').annotate(total=Count('id')).order_by('cable_type')
    
    # Equipment Summary - CPAN and MAAN separately
    cpan_nodes = Equipment.objects.filter(equipment_type__startswith='CPAN').values('name', 'te__name', 'total_ports')
    maan_nodes = Equipment.objects.filter(equipment_type__startswith='MAAN').values('name', 'equipment_type', 'te__name', 'total_ports')
    
    if division:
        cpan_nodes = cpan_nodes.filter(te__nwo=division)
        maan_nodes = maan_nodes.filter(te__nwo=division)
    
    equipment_queryset = Equipment.objects.all()
    if division:
        equipment_queryset = equipment_queryset.filter(te__nwo=division)
    equipment_summary = equipment_queryset.values('equipment_type').annotate(total=Count('id')).order_by('equipment_type')
    
    # Circuit Summary by Type
    circuit_queryset = EBCircuit.objects.all()
    if division:
        circuit_queryset = circuit_queryset.filter(te__nwo=division)
    circuit_type_summary = circuit_queryset.values('circuit_type').annotate(total=Count('id')).order_by('circuit_type')
    
    # JB Summary - OH and UG separately
    jb_queryset = JunctionBox.objects.all()
    if division:
        jb_queryset = jb_queryset.filter(te__nwo=division)
    oh_jbs = jb_queryset.filter(jb_type='OH').count()
    ug_jbs = jb_queryset.filter(jb_type='UG').count()

    if request.user.is_superuser:
        divisions = list(NWO.objects.all().order_by('name'))
    else:
        divisions = [division] if division else []

    return render(request, 'inventory/analytics.html', {
        'divisions': divisions,
        'selected_division': division,
        'bts_list': bts_list,
        'cable_summary': cable_summary,
        'oh_cables': oh_cables,
        'ug_cables': ug_cables,
        'equipment_summary': equipment_summary,
        'cpan_nodes': cpan_nodes,
        'maan_nodes': maan_nodes,
        'cpan_count': cpan_nodes.count(),
        'maan_count': maan_nodes.count(),
        'circuit_type_summary': circuit_type_summary,
        'oh_jbs': oh_jbs,
        'ug_jbs': ug_jbs,
        # New Fault Tracking Context
        'division_mttr': division_mttr,
        'team_performance': team_performance,
        'repeated_locations': repeated_locations,
        'fault_type_trends': fault_type_trends,
        'severity_dist': severity_dist,
    })

@login_required
def export_analytics(request):
    division = None
    if request.user.is_superuser:
        division_id = request.GET.get('division')
        if division_id:
            division = NWO.objects.filter(id=division_id).first()
    else:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            division = request.user.profile.division

    bts_list = MobileBTS.objects.select_related('maan_node', 'maan_node__te').all()
    if division:
        bts_list = bts_list.filter(maan_node__te__nwo=division)

    oh_cables = Cable.objects.filter(mode='OH').annotate(
        total_fibers=Count('fibers'),
        used_fibers=Count('fibers', filter=Q(fibers__is_used=True)),
        free_fibers=Count('fibers', filter=Q(fibers__is_used=False))
    ).select_related('te')

    ug_cables = Cable.objects.filter(mode='UG').annotate(
        total_fibers=Count('fibers'),
        used_fibers=Count('fibers', filter=Q(fibers__is_used=True)),
        free_fibers=Count('fibers', filter=Q(fibers__is_used=False))
    ).select_related('te')

    cable_queryset = Cable.objects.all()
    equipment_queryset = Equipment.objects.all()
    circuit_queryset = EBCircuit.objects.all()

    if division:
        oh_cables = oh_cables.filter(te__nwo=division)
        ug_cables = ug_cables.filter(te__nwo=division)
        cable_queryset = cable_queryset.filter(te__nwo=division)
        equipment_queryset = equipment_queryset.filter(te__nwo=division)
        circuit_queryset = circuit_queryset.filter(te__nwo=division)

    cable_summary = cable_queryset.values('cable_type', 'mode').annotate(total=Count('id')).order_by('cable_type', 'mode')
    equipment_summary = equipment_queryset.values('equipment_type').annotate(total=Count('id')).order_by('equipment_type')
    circuit_type_summary = circuit_queryset.values('circuit_type').annotate(total=Count('id')).order_by('circuit_type')

    cpan_nodes = Equipment.objects.filter(equipment_type__startswith='CPAN')
    maan_nodes = Equipment.objects.filter(equipment_type__startswith='MAAN')
    if division:
        cpan_nodes = cpan_nodes.filter(te__nwo=division)
        maan_nodes = maan_nodes.filter(te__nwo=division)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    division_label = division.name if division else "ALL"
    ws.append(["Division", division_label])
    ws.append([])
    ws.append(["Metric", "Value"])

    metrics = [
        ("Total BTS", bts_list.count()),
        ("Cable Types", cable_summary.count()),
        ("Total Cables", cable_queryset.count()),
        ("Total Equipment", equipment_queryset.count()),
        ("Total Circuits", circuit_queryset.count()),
        ("CPAN Nodes", cpan_nodes.count()),
        ("MAAN Nodes", maan_nodes.count()),
    ]
    for k, v in metrics:
        ws.append([k, v])

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    ws_cable_types = wb.create_sheet("Cable Types")
    ws_cable_types.append(["Cable Type", "Mode", "Total"])
    for row in cable_summary:
        ws_cable_types.append([row.get("cable_type"), row.get("mode"), row.get("total")])
    ws_cable_types.freeze_panes = "A2"

    ws_oh = wb.create_sheet("Cables OH")
    ws_oh.append(["Cable Name", "TE", "Type", "Total Fibers", "Used", "Free"])
    for c in oh_cables.order_by("te__name", "name"):
        ws_oh.append([c.name, (c.te.name or "").upper(), c.cable_type, c.total_fibers, c.used_fibers, c.free_fibers])
    ws_oh.freeze_panes = "A2"

    ws_ug = wb.create_sheet("Cables UG")
    ws_ug.append(["Cable Name", "TE", "Type", "Total Fibers", "Used", "Free"])
    for c in ug_cables.order_by("te__name", "name"):
        ws_ug.append([c.name, (c.te.name or "").upper(), c.cable_type, c.total_fibers, c.used_fibers, c.free_fibers])
    ws_ug.freeze_panes = "A2"

    ws_equipment_types = wb.create_sheet("Equipment Types")
    ws_equipment_types.append(["Equipment Type", "Total"])
    for row in equipment_summary:
        ws_equipment_types.append([row.get("equipment_type"), row.get("total")])
    ws_equipment_types.freeze_panes = "A2"

    ws_cpan = wb.create_sheet("CPAN Nodes")
    ws_cpan.append(["Node Name", "TE", "Total Ports"])
    for eq in cpan_nodes.select_related("te").order_by("te__name", "name"):
        ws_cpan.append([eq.name, (eq.te.name or "").upper(), eq.total_ports])
    ws_cpan.freeze_panes = "A2"

    ws_maan = wb.create_sheet("MAAN Nodes")
    ws_maan.append(["Node Name", "Equipment Type", "TE", "Total Ports"])
    for eq in maan_nodes.select_related("te").order_by("te__name", "name"):
        ws_maan.append([eq.name, eq.equipment_type, (eq.te.name or "").upper(), eq.total_ports])
    ws_maan.freeze_panes = "A2"

    ws_circuits = wb.create_sheet("Circuit Types")
    ws_circuits.append(["Circuit Type", "Total"])
    for row in circuit_type_summary:
        ws_circuits.append([row.get("circuit_type"), row.get("total")])
    ws_circuits.freeze_panes = "A2"

    ws_bts = wb.create_sheet("BTS")
    ws_bts.append(["RP ID", "BTS Name", "TE", "Place", "Has CEF 12T", "Ring", "Latitude", "Longitude"])
    for b in bts_list.order_by("rp_id"):
        te_name = ""
        if b.maan_node and b.maan_node.te:
            te_name = (b.maan_node.te.name or "").upper()
        ws_bts.append([b.rp_id, b.bts_name, te_name, b.place_name or "", "Y" if b.has_cef_12t else "N", "Y" if b.is_ring else "N", b.latitude, b.longitude])
    ws_bts.freeze_panes = "A2"

    for sheet in wb.worksheets:
        max_col = sheet.max_column
        for col_idx in range(1, max_col + 1):
            column_letter = get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = max(sheet.column_dimensions[column_letter].width or 10, 14)

    filename = f"analytics_report_{division_label.lower().replace(' ', '_')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_cables(request):
    cables = Cable.objects.all()
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            cables = cables.filter(te__nwo=request.user.profile.division)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cables Report"
    
    columns = ['Name', 'Cable Type', 'Mode', 'TE Name', 'Remarks']
    ws.append(columns)
    
    for cable in cables:
        ws.append([
            cable.name,
            cable.cable_type,
            cable.mode,
            cable.te.name if cable.te else "N/A",
            cable.remarks or ""
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cables_report.xlsx"'
    wb.save(response)
    return response

@login_required
def export_equipment(request):
    equipment = Equipment.objects.all()
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            equipment = equipment.filter(te__nwo=request.user.profile.division)
            
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Report"
    
    columns = ['Name', 'Equipment Type', 'Total Ports', 'TE Name', 'Remarks']
    ws.append(columns)
    
    for eq in equipment:
        ws.append([
            eq.name,
            eq.equipment_type,
            eq.total_ports,
            eq.te.name if eq.te else "N/A",
            eq.remarks or ""
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="equipment_report.xlsx"'
    wb.save(response)
    return response

@login_required
def export_bts(request):
    queryset = MobileBTS.objects.select_related('maan_node', 'maan_node__te').all()
    division = None
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
        division = request.user.profile.division
        queryset = queryset.filter(maan_node__te__nwo=division)

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(Q(rp_id__icontains=search) | Q(bts_name__icontains=search))

    is_ring = request.GET.get('is_ring')
    if is_ring:
        queryset = queryset.filter(is_ring=is_ring == 'true')

    wb = Workbook()
    ws = wb.active
    ws.title = "BTS Report"

    columns = [
        'RP ID', 'BTS Name', 'MAAN Node', 'TE', 'Place Name',
        'Latitude', 'Longitude', 'Has CEF 12T', 'Is Ring', 'ERPS Image URL',
        'P2 Circuit', 'P2 System End', 'P2 Cable',
        'P3 Circuit', 'P3 System End', 'P3 Cable',
        'P4 Circuit', 'P4 System End', 'P4 Cable',
        'P5 Circuit', 'P5 System End', 'P5 Cable',
    ]
    ws.append(columns)

    for bts in queryset.order_by('rp_id'):
        ports = bts.cef_ports_data or {}
        def port_value(port_key, field_key):
            data = ports.get(port_key, {})
            return data.get(field_key, '') if isinstance(data, dict) else ''

        te_name = ''
        maan_node_name = ''
        if bts.maan_node:
            maan_node_name = bts.maan_node.name or ''
            if bts.maan_node.te:
                te_name = bts.maan_node.te.name or ''

        erps_url = ''
        if getattr(bts, 'erps_image', None):
            try:
                erps_url = bts.erps_image.url
            except Exception:
                erps_url = ''

        ws.append([
            bts.rp_id,
            bts.bts_name,
            maan_node_name,
            te_name,
            bts.place_name or '',
            float(bts.latitude) if bts.latitude is not None else '',
            float(bts.longitude) if bts.longitude is not None else '',
            'YES' if bts.has_cef_12t else 'NO',
            'YES' if bts.is_ring else 'NO',
            erps_url,
            port_value('p2', 'circuit'),
            port_value('p2', 'system_end'),
            port_value('p2', 'cable'),
            port_value('p3', 'circuit'),
            port_value('p3', 'system_end'),
            port_value('p3', 'cable'),
            port_value('p4', 'circuit'),
            port_value('p4', 'system_end'),
            port_value('p4', 'cable'),
            port_value('p5', 'circuit'),
            port_value('p5', 'system_end'),
            port_value('p5', 'cable'),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    division_label = division.name if division else 'all'
    filename = f"bts_report_{division_label.lower().replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_ftth(request):
    queryset = FTTH.objects.select_related('division', 'te').all()
    division = None
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
        division = request.user.profile.division
        queryset = queryset.filter(division=division)

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(
            Q(customer_name__icontains=search) |
            Q(landline_number__icontains=search) |
            Q(olt_name__icontains=search)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "FTTH Report"

    columns = [
        'Customer Name', 'Landline Number', 'Optical Power (dB)', 'OLT Name', 'Port Number',
        'Division', 'TE', 'Latitude', 'Longitude', 'Created At'
    ]
    ws.append(columns)

    for ftth in queryset.order_by('-created_at'):
        division_name = ''
        if ftth.division:
            try:
                division_name = ftth.division.get_name_display()
            except Exception:
                division_name = ftth.division.name

        te_name = ftth.te.name if ftth.te else ''

        ws.append([
            ftth.customer_name,
            ftth.landline_number,
            float(ftth.optical_power) if ftth.optical_power is not None else '',
            ftth.olt_name,
            ftth.port_number,
            division_name,
            te_name,
            float(ftth.latitude) if ftth.latitude is not None else '',
            float(ftth.longitude) if ftth.longitude is not None else '',
            ftth.created_at,
        ])

    ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width or 10, 16)

    division_label = division.name if division else 'all'
    filename = f"ftth_report_{division_label.lower().replace(' ', '_')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_circuits(request):
    queryset = EBCircuit.objects.select_related('te', 'te__nwo', 'cable', 'equipment').all()
    division = None
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
        division = request.user.profile.division
        queryset = queryset.filter(te__nwo=division)

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(client_name__icontains=search)

    circuit_type = request.GET.get('circuit_type')
    if circuit_type:
        queryset = queryset.filter(circuit_type=circuit_type)

    te_id = request.GET.get('te')
    if te_id:
        queryset = queryset.filter(te_id=te_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "EB Circuits"

    columns = [
        'Client Name', 'Type', 'Bandwidth', 'End Node', 'Fiber Mode',
        'Division', 'TE', 'Cable', 'Equipment',
        'Customer Premise Location', 'OTDR Distance', 'Latitude', 'Longitude', 'Remarks'
    ]
    ws.append(columns)

    for c in queryset.order_by('te__name', 'client_name'):
        division_name = c.te.nwo.get_name_display() if c.te and c.te.nwo else ''
        te_name = c.te.name if c.te else ''
        ws.append([
            c.client_name,
            c.circuit_type,
            c.bandwidth,
            c.customer_end_node,
            c.fiber_mode,
            division_name,
            te_name,
            c.cable.name if c.cable else '',
            c.equipment.name if c.equipment else '',
            c.customer_premise_location or '',
            c.otdr_distance or '',
            c.latitude or '',
            c.longitude or '',
            c.remarks or '',
        ])

    ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width or 10, 16)

    division_label = division.name if division else 'all'
    filename = f"eb_circuits_{division_label.lower().replace(' ', '_')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response

@login_required
def bulk_upload(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        upload_type = request.POST.get('upload_type')
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            header_map = {name: i for i, name in enumerate(headers)}
            
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            def get_value(row, *keys, required=False, default=None):
                for key in keys:
                    idx = header_map.get(key)
                    if idx is not None:
                        return row[idx]
                if required:
                    raise KeyError(f"Missing required column: {keys[0]}")
                return default

            def parse_bool(val):
                if val is None:
                    return False
                if isinstance(val, bool):
                    return val
                text = str(val).strip().lower()
                return text in {"y", "yes", "true", "1", "t"}

            created_count = 0
            skipped_count = 0
            row_errors = []
            
            if upload_type == 'CIRCUIT':
                for i, row in enumerate(rows, start=2):
                    try:
                        te_name = get_value(row, 'TE', 'TE Name', required=True)
                        te_queryset = TelephoneExchange.objects.filter(name=te_name)
                        if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                            te_queryset = te_queryset.filter(nwo=request.user.profile.division)
                        te = te_queryset.first()
                        if not te:
                            skipped_count += 1
                            continue

                        EBCircuit.objects.create(
                            client_name=get_value(row, 'Client Name', required=True),
                            circuit_type=get_value(row, 'Type', required=True),
                            bandwidth=get_value(row, 'Bandwidth', required=True),
                            customer_end_node=get_value(row, 'End Node', required=True),
                            fiber_mode=get_value(row, 'Fiber Mode', required=True),
                            te=te,
                            remarks=get_value(row, 'Remarks', default='Bulk Uploaded') or 'Bulk Uploaded'
                        )
                        created_count += 1
                    except IntegrityError:
                        skipped_count += 1
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                if row_errors:
                    messages.warning(request, f"EB Circuits uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"EB Circuits uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")
            
            elif upload_type == 'CABLE':
                for i, row in enumerate(rows, start=2):
                    try:
                        te_name = get_value(row, 'TE', 'TE Name', required=True)
                        te_queryset = TelephoneExchange.objects.filter(name=te_name)
                        if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                            te_queryset = te_queryset.filter(nwo=request.user.profile.division)
                        te = te_queryset.first()
                        if not te:
                            skipped_count += 1
                            continue

                        Cable.objects.create(
                            name=get_value(row, 'Cable Name', required=True),
                            cable_type=get_value(row, 'Type', required=True),
                            fiber_count=get_value(row, 'Fiber Count', required=True),
                            mode=get_value(row, 'Mode', required=True),
                            te=te,
                            remarks=get_value(row, 'Remarks', default='Bulk Uploaded') or 'Bulk Uploaded'
                        )
                        created_count += 1
                    except IntegrityError:
                        skipped_count += 1
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                if row_errors:
                    messages.warning(request, f"Cables uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"Cables uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")

            elif upload_type == 'EQUIPMENT':
                for i, row in enumerate(rows, start=2):
                    try:
                        te_name = get_value(row, 'TE', 'TE Name', required=True)
                        te_queryset = TelephoneExchange.objects.filter(name=te_name)
                        if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                            te_queryset = te_queryset.filter(nwo=request.user.profile.division)
                        te = te_queryset.first()
                        if not te:
                            skipped_count += 1
                            continue

                        Equipment.objects.create(
                            name=get_value(row, 'Equipment Name', required=True),
                            equipment_type=get_value(row, 'Type', required=True),
                            total_ports=get_value(row, 'Total Ports', required=True),
                            te=te,
                            remarks=get_value(row, 'Remarks', default='Bulk Uploaded') or 'Bulk Uploaded'
                        )
                        created_count += 1
                    except IntegrityError:
                        skipped_count += 1
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                if row_errors:
                    messages.warning(request, f"Equipment uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"Equipment uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")

            elif upload_type == 'BTS':
                for i, row in enumerate(rows, start=2):
                    try:
                        rp_id = str(get_value(row, 'RP ID', required=True)).strip()
                        bts_name = get_value(row, 'BTS Name', required=True)
                        place_name = get_value(row, 'Place Name')
                        latitude = get_value(row, 'Latitude')
                        longitude = get_value(row, 'Longitude')
                        has_cef_12t = parse_bool(get_value(row, 'Has CEF 12T', default=False))
                        is_ring = parse_bool(get_value(row, 'Is Ring', default=False))

                        ports_data = {}
                        for port in ['P2', 'P3', 'P4', 'P5']:
                            ports_data[port.lower()] = {
                                'circuit': get_value(row, f'{port} Circuit', default='') or '',
                                'system_end': get_value(row, f'{port} System End', default='') or '',
                                'cable': get_value(row, f'{port} Cable', default='') or '',
                            }

                        MobileBTS.objects.create(
                            rp_id=rp_id,
                            bts_name=bts_name,
                            place_name=place_name,
                            latitude=Decimal(str(latitude)) if latitude not in (None, '') else None,
                            longitude=Decimal(str(longitude)) if longitude not in (None, '') else None,
                            has_cef_12t=has_cef_12t,
                            is_ring=is_ring,
                            cef_ports_data=ports_data,
                        )
                        created_count += 1
                    except IntegrityError:
                        skipped_count += 1
                    except (InvalidOperation, ValueError) as e:
                        row_errors.append(f"Row {i}: Invalid numeric value ({e})")
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                if row_errors:
                    messages.warning(request, f"BTS uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"BTS uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")

            elif upload_type == 'FTTH':
                user_division = None
                if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                    user_division = request.user.profile.division

                for i, row in enumerate(rows, start=2):
                    try:
                        customer_name = get_value(row, 'Customer Name', required=True)
                        landline_number = str(get_value(row, 'Landline Number', required=True)).strip()
                        optical_power_raw = get_value(row, 'Optical Power', required=True)
                        olt_name = get_value(row, 'OLT Name', required=True)
                        port_number = int(get_value(row, 'Port Number', required=True))
                        division_name = get_value(row, 'Division')
                        te_name = get_value(row, 'TE', 'TE Name')
                        latitude = get_value(row, 'Latitude')
                        longitude = get_value(row, 'Longitude')

                        optical_power_str = str(optical_power_raw).strip()
                        optical_power_value = Decimal(optical_power_str.replace('dB', '').replace('DB', '').strip())

                        division = user_division
                        if division is None and division_name not in (None, ''):
                            division = NWO.objects.filter(name=str(division_name).strip()).first()

                        te = None
                        if te_name not in (None, ''):
                            te_queryset = TelephoneExchange.objects.filter(name=str(te_name).strip())
                            if division is not None:
                                te_queryset = te_queryset.filter(nwo=division)
                            te = te_queryset.first()
                            if division is None and te is not None:
                                division = te.nwo

                        FTTH.objects.create(
                            customer_name=customer_name,
                            landline_number=landline_number,
                            optical_power=optical_power_value,
                            olt_name=olt_name,
                            port_number=port_number,
                            division=division,
                            te=te,
                            latitude=Decimal(str(latitude)) if latitude not in (None, '') else None,
                            longitude=Decimal(str(longitude)) if longitude not in (None, '') else None,
                        )
                        created_count += 1
                    except IntegrityError:
                        skipped_count += 1
                    except (InvalidOperation, ValueError) as e:
                        row_errors.append(f"Row {i}: Invalid numeric value ({e})")
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                if row_errors:
                    messages.warning(request, f"FTTH uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"FTTH uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")
            
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            
    return render(request, 'inventory/bulk_upload.html')

@login_required
def download_template(request):
    category = request.GET.get('category', 'CIRCUIT')
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{category} Template"
    
    if category == 'CIRCUIT':
        columns = ['Client Name', 'Type', 'Bandwidth', 'End Node', 'Fiber Mode', 'TE', 'Remarks']
        data = ['Reliance JIO', 'ILL', '100 Mbps', 'CPE', 'DUAL', 'Panambilly Nagar', 'Sample circuit']
    elif category == 'CABLE':
        columns = ['Cable Name', 'Type', 'Fiber Count', 'Mode', 'TE', 'Remarks']
        data = ['PN-CSR-48F-01', '48F', 48, 'UG', 'Panambilly Nagar', 'Sample cable']
    elif category == 'EQUIPMENT':
        columns = ['Equipment Name', 'Type', 'Total Ports', 'TE', 'Remarks']
        data = ['PN-CPAN-01', 'CPAN_B', 24, 'Panambilly Nagar', 'Sample equipment']
    elif category == 'BTS':
        columns = [
            'RP ID', 'BTS Name', 'Place Name', 'Latitude', 'Longitude', 'Has CEF 12T', 'Is Ring',
            'P2 Circuit', 'P2 System End', 'P2 Cable',
            'P3 Circuit', 'P3 System End', 'P3 Cable',
            'P4 Circuit', 'P4 System End', 'P4 Cable',
            'P5 Circuit', 'P5 System End', 'P5 Cable',
        ]
        data = [
            'RP-0001', 'BTS Example Site', 'Sample Location', 9.931233, 76.267303, 'Y', 'N',
            'CIR-001', 'SYSTEM-A', 'CABLE-001',
            '', '', '',
            '', '', '',
            '', '', '',
        ]
    elif category == 'FTTH':
        columns = ['Customer Name', 'Landline Number', 'Optical Power', 'OLT Name', 'Port Number', 'Division', 'TE', 'Latitude', 'Longitude']
        data = ['Customer A', '0484XXXXXXX', '-23.45', 'OLT-1', 1, 'NWO KOCHI', 'Panambilly Nagar', 9.931233, 76.267303]
    else:
        columns = ['Client Name', 'Type', 'Bandwidth', 'End Node', 'Fiber Mode', 'TE', 'Remarks']
        data = ['Reliance JIO', 'ILL', '100 Mbps', 'CPE', 'DUAL', 'Panambilly Nagar', 'Sample circuit']
        
    ws.append(columns)
    ws.append(data)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{category.lower()}_template.xlsx"'
    wb.save(response)
    return response

# OH Maintenance Module Views

class OHMaintenanceListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = OHMaintenanceEntry
    template_name = 'inventory/oh_maintenance_list.html'
    context_object_name = 'entries'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        # Search filters
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(team_name__icontains(q)) |
                Q(work_order_no__icontains(q)) |
                Q(location__icontains(q)) |
                Q(route_name__icontains(q))
            )
        
        # Date range filter
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date:
            queryset = queryset.filter(maintenance_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(maintenance_date__lte=end_date)
            
        # TE filter
        te_id = self.request.GET.get('te')
        if te_id:
            queryset = queryset.filter(te_id=te_id)
            
        return queryset

class OHMaintenanceCreateView(LoginRequiredMixin, DivisionRequiredMixin, CreateView):
    model = OHMaintenanceEntry
    form_class = OHMaintenanceEntryForm
    template_name = 'inventory/oh_maintenance_form.html'
    success_url = reverse_lazy('oh_maintenance_list')

    def get_initial(self):
        initial = super().get_initial()
        te_id = self.request.GET.get('te')
        if te_id:
            initial['te'] = te_id
        return initial

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['activities'] = OHMaintenanceActivityFormSet(self.request.POST)
        else:
            data['activities'] = OHMaintenanceActivityFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        activities = context['activities']
        
        # Auto-fetch division from profile
        if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
            form.instance.division = self.request.user.profile.division
        
        form.instance.created_by = self.request.user
        
        if activities.is_valid():
            self.object = form.save()
            activities.instance = self.object
            activities.save()
            messages.success(self.request, "Maintenance entry created successfully.")
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class OHMaintenanceUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = OHMaintenanceEntry
    form_class = OHMaintenanceEntryForm
    template_name = 'inventory/oh_maintenance_form.html'
    success_url = reverse_lazy('oh_maintenance_list')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['activities'] = OHMaintenanceActivityFormSet(self.request.POST, instance=self.object)
        else:
            data['activities'] = OHMaintenanceActivityFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        activities = context['activities']
        
        form.instance.modified_by = self.request.user
        
        if activities.is_valid():
            self.object = form.save()
            activities.instance = self.object
            activities.save()
            messages.success(self.request, "Maintenance entry updated successfully.")
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class OHMaintenanceDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = OHMaintenanceEntry
    template_name = 'inventory/oh_maintenance_detail.html'
    context_object_name = 'entry'

class OHMaintenanceDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = OHMaintenanceEntry
    template_name = 'inventory/oh_maintenance_confirm_delete.html'
    success_url = reverse_lazy('oh_maintenance_list')

@login_required
def get_activity_rates(request):
    rates = OHMaintenanceRateMaster.objects.filter(is_active=True).values(
        'activity_type', 'unit_type', 'unit_rate'
    )
    return JsonResponse(list(rates), safe=False)

class OHMaintenanceRateMasterListView(LoginRequiredMixin, ListView):
    model = OHMaintenanceRateMaster
    template_name = 'inventory/oh_rate_master_list.html'
    context_object_name = 'rates'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, "Only admins can access Rate Master.")
            return redirect('oh_maintenance_list')
        return super().dispatch(request, *args, **kwargs)

class OHMaintenanceRateMasterCreateView(LoginRequiredMixin, CreateView):
    model = OHMaintenanceRateMaster
    fields = ['activity_type', 'unit_type', 'unit_rate', 'effective_from', 'effective_to', 'is_active']
    template_name = 'inventory/oh_rate_master_form.html'
    success_url = reverse_lazy('oh_rate_master_list')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, "Only admins can access Rate Master.")
            return redirect('oh_maintenance_list')
        return super().dispatch(request, *args, **kwargs)

@login_required
def oh_maintenance_bill(request):
    division = None
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            division = request.user.profile.division
    else:
        division_id = request.GET.get('division')
        if division_id:
            division = NWO.objects.filter(id=division_id).first()

    month = request.GET.get('month', timezone.now().month)
    year = request.GET.get('year', timezone.now().year)
    
    entries = OHMaintenanceEntry.objects.filter(
        status='Approved',
        maintenance_date__month=month,
        maintenance_date__year=year
    )
    if division:
        entries = entries.filter(division=division)
        
    # Aggregate by activity type
    bill_data = entries.values('activities__activity_type', 'activities__unit_type', 'activities__unit_rate').annotate(
        total_quantity=Sum('activities__quantity'),
        total_amount=Sum('activities__amount')
    ).order_by('activities__activity_type')
    
    total_bill_amount = sum(item['total_amount'] or 0 for item in bill_data)

    context = {
        'bill_data': bill_data,
        'total_bill_amount': total_bill_amount,
        'month': month,
        'year': year,
        'division': division,
        'divisions': NWO.objects.all() if request.user.is_superuser else None
    }
    return render(request, 'inventory/oh_maintenance_bill.html', context)

@login_required
def oh_maintenance_dashboard(request):
    from django.db.models import F, ExpressionWrapper, DurationField, Avg, Count, Sum
    from django.utils import timezone
    from datetime import timedelta
    
    division = None
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            division = request.user.profile.division
    else:
        div_id = request.GET.get('division')
        if div_id:
            division = NWO.objects.filter(id=div_id).first()
            
    entries = OHMaintenanceEntry.objects.all()
    if division:
        entries = entries.filter(division=division)
    
    # Fault Tracking Stats
    now = timezone.now()
    first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calculate durations for MTTR
    entries_with_duration = entries.filter(
        fault_occurrence_datetime__isnull=False,
        fault_clearance_datetime__isnull=False
    ).annotate(
        duration=ExpressionWrapper(
            F('fault_clearance_datetime') - F('fault_occurrence_datetime'),
            output_field=DurationField()
        )
    )
    
    avg_mttr = entries_with_duration.aggregate(avg=Avg('duration'))['avg']
    
    stats = {
        'total_entries': entries.count(),
        'pending_verification': entries.filter(status='Submitted').count(),
        'pending_approval': entries.filter(status='Verified').count(),
        'total_cost_this_month': entries.filter(
            maintenance_date__month=now.month,
            maintenance_date__year=now.year
        ).aggregate(total=Sum('activities__amount'))['total'] or 0,
        
        # New Fault Tracking Stats
        'total_faults_this_month': entries.filter(
            fault_occurrence_datetime__gte=first_day_of_month
        ).count(),
        'avg_clearance_time': avg_mttr,
        'critical_fault_count': entries.filter(fault_severity='Critical').count(),
        'pending_fault_closures': entries.filter(
            fault_occurrence_datetime__isnull=False,
            fault_clearance_datetime__isnull=True
        ).count(),
    }
    
    # Division-wise MTTR for Graph
    division_mttr = entries_with_duration.values('division__name').annotate(
        avg_duration=Avg('duration')
    ).order_by('division__name')
    
    # Convert timedelta to hours for display in graph
    for item in division_mttr:
        if item['avg_duration']:
            item['avg_hours'] = item['avg_duration'].total_seconds() / 3600
        else:
            item['avg_hours'] = 0

    # Repeated Fault Hotspot Locations
    hotspots = entries.values('location').annotate(
        fault_count=Count('id')
    ).filter(fault_count__gt=1).order_by('-fault_count')[:10]
    
    # Activity-wise summary for chart
    activity_summary = OHMaintenanceActivity.objects.filter(entry__in=entries).values('activity_type').annotate(
        total_qty=Sum('quantity')
    ).order_by('-total_qty')

    context = {
        'stats': stats,
        'activity_summary': activity_summary,
        'recent_entries': entries[:10],
        'division_mttr': division_mttr,
        'hotspots': hotspots,
        'divisions': NWO.objects.all() if request.user.is_superuser else None,
        'selected_division': division,
    }
    return render(request, 'inventory/oh_maintenance_dashboard.html', context)
